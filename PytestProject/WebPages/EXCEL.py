from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By

driver = webdriver.Firefox()

driver.maximize_window()
driver.implicitly_wait(10)

# Assuming your Excel file is in the same directory as your script
workbook = load_workbook("Book1.xlsx")
sheet = workbook["Sheet1"]

row_count = sheet.max_row

for curr_row in range(2, row_count + 1):  # Start from 2 to skip the header row
    user_value = sheet.cell(curr_row, 1).value
    pass_value = sheet.cell(curr_row, 2).value

    # Navigate to your web page
    driver.get("http://www.isiri.co/hrm/")

    # Find the username and password fields and input the values
    username_input = driver.find_element(By.ID, "txtUsername")
    password_input = driver.find_element(By.ID, "txtPassword")

    username_input.send_keys(user_value)
    password_input.send_keys(pass_value)

    # Perform login or other actions as needed
    # For example, you can add a click action for a login button if it exists
    # login_button = driver.find_element(By.ID, "loginButtonId")
    # login_button.click()

    # Add any other actions you want to perform on the page after login

# Close the browser after processing all rows
driver.quit()
