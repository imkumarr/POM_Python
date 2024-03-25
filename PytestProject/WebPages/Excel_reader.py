import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By

from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


driver = webdriver.Firefox()

driver.maximize_window()
driver.implicitly_wait(10)
wait = WebDriverWait(driver, 10)
driver.get("http://www.isiri.co/hrm/")
driver.find_element(By.ID, "txtUsername").send_keys("kumar")
driver.find_element(By.ID, "txtPassword").send_keys("iSiriTech!23")

driver.find_element(By.ID, "btnLogin").click()

workbook = load_workbook("Excel_Data.xlsx")
sheet = workbook["Sheet1"]

row_count = sheet.max_row

for curr_row in range(2, row_count + 1):  # Start from 2 to skip the header row
    First_Name = sheet.cell(curr_row, 1).value
    Middle_Name = sheet.cell(curr_row, 2).value
    Last_Name = sheet.cell(curr_row, 3).value
    Employee_Id = sheet.cell(curr_row, 4).value
    Photograph=sheet.cell(curr_row, 5).value
    Other_Id = sheet.cell(curr_row, 6).value
    Dri_Lic = sheet.cell(curr_row, 7).value
    LExpD = sheet.cell(curr_row, 8).value
    LExpM = sheet.cell(curr_row, 9).value
    LExpY = sheet.cell(curr_row, 10).value
    Gender = sheet.cell(curr_row, 11).value
    Marital_Stu= sheet.cell(curr_row, 12).value
    Nationality= sheet.cell(curr_row, 13).value
    DBD= sheet.cell(curr_row, 14).value
    DBM= sheet.cell(curr_row, 15).value
    DBY= sheet.cell(curr_row, 16).value

    print( First_Name,Middle_Name,Last_Name,Employee_Id,Photograph,Other_Id,Dri_Lic,LExpD,LExpM,LExpY,Gender,Marital_Stu,Nationality,DBD,DBM,DBY)

    # time.sleep(6)
    driver.find_element(By.ID, "menu_pim_viewPimModule").click()
    driver.find_element(By.ID, "menu_pim_addEmployee").click()

    first_name=driver.find_element(By.NAME, "firstName")
    # first_name.clear()
    last_name=driver.find_element(By.NAME, "lastName")
    id=driver.find_element(By.NAME, "employeeId")
    # id.clear()
    ph=driver.find_element(By.XPATH, "//input[@id='photofile']")

    first_name.clear()
    first_name.send_keys(First_Name)
    last_name.clear()
    last_name.send_keys(Last_Name)
    id.clear()
    id.send_keys(Employee_Id)
    ph.send_keys(Photograph)

    driver.find_element(By.XPATH, "//input[@id='btnSave']").click()

    # menu= wait.until(EC.element_to_be_clickable((By.ID,"menu_pim_viewPimModule")))
    # menu.click()
    # driver.find_element(By.ID,"empsearch_id").clear()
    # driver.find_element(By.ID,"empsearch_id").send_keys(Employee_Id)
    # driver.find_element(By.NAME,"_search").click()
    # driver.find_element(By.LINK_TEXT, "Employee_Id").click()
    driver.find_element(By.ID,"btnSave").click()
    #

    other_ids=driver.find_element(By.ID,"personal_txtOtherID")
    other_ids.clear()
    other_ids.send_keys(Other_Id)
    Lic_no=driver.find_element(By.NAME,"personal[txtLicenNo]")
    Lic_no.clear()
    Lic_no.send_keys(Dri_Lic)
    # def Lic_Exp(option, value):
    #     year=Select(option)
    #     year.select_by_visible_text(value)
    def select_dropdown(options, value):
        for ele in options:
            if ele.text==value:
                ele.click()
                break

    driver.find_element(By.NAME,"personal[txtLicExpDate]").click()
    LEm=driver.find_elements(By.XPATH,"//select[@class='ui-datepicker-month']/option")
    select_dropdown(LEm,LExpM)
    print(LExpM)
    LEy=driver.find_elements(By.XPATH,"//select[@class='ui-datepicker-year']/option")
    select_dropdown(LEy,LExpY)
    LEd=driver.find_elements(By.XPATH,"//td[@data-handler='selectDay']/a")
    select_dropdown(LEd,LExpD)
    LEd[0].click()


    time.sleep(8)

    option = '1'
    male_checked = driver.execute_script("return document.getElementById('personal_optGender_1').checked;")
    female_checked = driver.execute_script("return document.getElementById('personal_optGender_2').checked;")

    if option == '1' and not male_checked:
        driver.execute_script("document.getElementById('personal_optGender_1').click()")
    elif option == '2' and not female_checked:
        driver.execute_script("document.getElementById('personal_optGender_2').click()")



    # gender=driver.find_elements(By.XPATH, "//ul[@class='radio_list']/li/input")
    # select_dropdown(gender, '1')
    mar_opt=driver.find_elements(By.XPATH, "//select[@id='personal_cmbMarital']/option")
    select_dropdown(mar_opt,Marital_Stu)
    mar_opt=driver.find_elements(By.XPATH, "//select[@id='personal_cmbNation']/option")
    select_dropdown(mar_opt,Nationality)

    # date of birth
    driver.find_element(By.ID,"personal_DOB").click()
    DM=driver.find_elements(By.XPATH,"//select[@class='ui-datepicker-month']/option")
    select_dropdown(DM,DBD)
    DY=driver.find_elements(By.XPATH,"//select[@class='ui-datepicker-year']/option")
    select_dropdown(DY,DBM)
    DD=driver.find_elements(By.XPATH,"//td[@data-handler='selectDay']/a")
    select_dropdown(DD,DBY)

    DD[0].click()




    driver.find_element(By.ID,"btnSave").click()
    driver.find_element(By.ID,"menu_pim_viewPimModule").click()
    search = driver.find_element(By.ID,"empsearch_id")
    search.clear()

    search.send_keys(Employee_Id )
    driver.find_element(By.NAME,"_search").click()
    delete=wait.until(EC.element_to_be_clickable((By.ID,"ohrmList_chkSelectAll")))
    delete.click()
    parentHandle = driver.current_window_handle
    driver.find_element(By.ID,"btnDelete").click()
    handles = driver.window_handles
    for handle in handles:
        if handle != parentHandle:
            driver.switch_to.window(handle)

    driver.find_element(By.ID,"dialogDeleteBtn").click()

driver.quit()
